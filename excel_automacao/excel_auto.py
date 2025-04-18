"""

from openpyxl import Workbook
objWorkbook = Workbook()
objWorksheet = objWorkbook.worksheets[0]
objWorksheet.cell(1,1,'This is my first task using Openpyxl')
objWorkbook.save("Learn_Excel_Automation.xlsx")
"""
"""
from openpyxl import load_workbook
objWorkbook = load_workbook("Learn_Excel_Automation.xlsx")
objWorksheet = objWorkbook.worksheets[0]
 
objWorksheet.cell(3,1,'This is my second task in openpyxl')
objWorkbook.save("Learn_Excel_Automation.xlsx")
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Criar uma nova planilha
objWorkbook = Workbook()
objWorksheet = objWorkbook.worksheets[0]

# Definir os cabeçalhos
objWorksheet.cell(1, 1, 'País')
objWorksheet.cell(1, 2, 'Capital')
objWorksheet.cell(1, 3, 'População (estimativa 2025)')

# Lista de países, capitais e população (estimativas arredondadas para 2025)
paises_capitais_populacao = [
    ('Afeganistão', 'Cabul', 42_000_000),
    ('África do Sul', 'Pretória', 63_000_000),  # Nota: África do Sul tem 3 capitais; Pretoria é a executiva
    ('Albânia', 'Tirana', 2_800_000),
    ('Alemanha', 'Berlim', 84_000_000),
    ('Andorra', 'Andorra la Vella', 80_000),
    ('Angola', 'Luanda', 37_000_000),
    ('Antígua e Barbuda', 'Saint John’s', 100_000),
    ('Arábia Saudita', 'Riade', 36_000_000),
    ('Argélia', 'Argel', 47_000_000),
    ('Argentina', 'Buenos Aires', 46_000_000),
    ('Armênia', 'Erevan', 3_000_000),
    ('Austrália', 'Camberra', 27_000_000),
    ('Áustria', 'Viena', 9_000_000),
    ('Azerbaijão', 'Baku', 10_000_000),
    ('Bahamas', 'Nassau', 400_000),
    ('Bahrein', 'Manama', 1_500_000),
    ('Bangladesh', 'Daca', 170_000_000),
    ('Barbados', 'Bridgetown', 300_000),
    ('Bélgica', 'Bruxelas', 12_000_000),
    ('Belize', 'Belmopan', 400_000),
    ('Benin', 'Porto-Novo', 14_000_000),
    ('Bielorrússia', 'Minsk', 9_300_000),
    ('Bolívia', 'La Paz', 12_000_000),  # La Paz é a sede do governo; Sucre é a capital constitucional
    ('Bósnia e Herzegovina', 'Sarajevo', 3_300_000),
    ('Botsuana', 'Gaborone', 2_700_000),
    ('Brasil', 'Brasília', 216_000_000),
    ('Brunei', 'Bandar Seri Begawan', 450_000),
    ('Bulgária', 'Sófia', 6_800_000),
    ('Burkina Faso', 'Ouagadougou', 23_000_000),
    ('Burundi', 'Gitega', 13_000_000),
    ('Butão', 'Thimbu', 800_000),
    ('Cabo Verde', 'Praia', 600_000),
    ('Camarões', 'Yaoundé', 29_000_000),
    ('Camboja', 'Phnom Penh', 17_000_000),
    ('Canadá', 'Ottawa', 41_000_000),
    ('Catar', 'Doha', 3_000_000),
    ('Cazaquistão', 'Astana', 20_000_000),
    ('Chade', 'N’Djamena', 18_000_000),
    ('Chile', 'Santiago', 20_000_000),
    ('China', 'Pequim', 1_425_000_000),
    ('Chipre', 'Nicósia', 1_300_000),
    ('Colômbia', 'Bogotá', 52_000_000),
    ('Comores', 'Moroni', 900_000),
    ('Congo (República do)', 'Brazzaville', 6_000_000),
    ('Coreia do Norte', 'Pyongyang', 26_000_000),
    ('Coreia do Sul', 'Seul', 51_000_000),
    ('Costa do Marfim', 'Yamoussoukro', 30_000_000),
    ('Costa Rica', 'San José', 5_200_000),
    ('Croácia', 'Zagreb', 4_000_000),
    ('Cuba', 'Havana', 11_000_000),
    ('Dinamarca', 'Copenhague', 6_000_000),
    ('Djibouti', 'Djibouti', 1_100_000),
    ('Dominica', 'Roseau', 70_000),
    ('Egito', 'Cairo', 110_000_000),
    ('El Salvador', 'San Salvador', 6_500_000),
    ('Emirados Árabes Unidos', 'Abu Dhabi', 10_000_000),
    ('Equador', 'Quito', 18_000_000),
    ('Eritreia', 'Asmara', 6_700_000),
    ('Eslováquia', 'Bratislava', 5_500_000),
    ('Eslovênia', 'Liubliana', 2_100_000),
    ('Espanha', 'Madri', 48_000_000),
    ('Estados Unidos', 'Washington, D.C.', 345_000_000),
    ('Estônia', 'Tallinn', 1_300_000),
    ('Etiópia', 'Adis Abeba', 126_000_000),
    ('Fiji', 'Suva', 900_000),
    ('Filipinas', 'Manila', 118_000_000),
    ('Finlândia', 'Helsinque', 5_600_000),
    ('França', 'Paris', 68_000_000),
    ('Gabão', 'Libreville', 2_400_000),
    ('Gâmbia', 'Banjul', 2_700_000),
    ('Gana', 'Acra', 34_000_000),
    ('Geórgia', 'Tbilisi', 3_700_000),
    ('Granada', 'Saint George’s', 100_000),
    ('Grécia', 'Atenas', 10_000_000),
    ('Guatemala', 'Cidade da Guatemala', 18_000_000),
    ('Guiana', 'Georgetown', 800_000),
    ('Guiné', 'Conakry', 14_000_000),
    ('Guiné Equatorial', 'Malabo', 1_700_000),
    ('Guiné-Bissau', 'Bissau', 2_100_000),
    ('Haiti', 'Porto Príncipe', 11_000_000),
    ('Honduras', 'Tegucigalpa', 10_000_000),
    ('Hungria', 'Budapeste', 9_600_000),
    ('Iêmen', 'Sana', 34_000_000),
    ('Índia', 'Nova Délhi', 1_450_000_000),
    ('Indonésia', 'Jacarta', 280_000_000),
    ('Irã', 'Teerã', 89_000_000),
    ('Iraque', 'Bagdá', 46_000_000),
    ('Irlanda', 'Dublin', 5_300_000),
    ('Islândia', 'Reykjavik', 400_000),
    ('Israel', 'Jerusalém', 9_300_000),  # Nota: Jerusalém é disputada; muitas embaixadas estão em Tel Aviv
    ('Itália', 'Roma', 59_000_000),
    ('Jamaica', 'Kingston', 2_800_000),
    ('Japão', 'Tóquio', 123_000_000),
    ('Jordânia', 'Amã', 11_000_000),
    ('Kiribati', 'Tarawa', 130_000),
    ('Kuwait', 'Cidade do Kuwait', 4_300_000),
    ('Laos', 'Vientiane', 7_700_000),
    ('Lesoto', 'Maseru', 2_300_000),
    ('Letônia', 'Riga', 1_800_000),
    ('Líbano', 'Beirute', 5_500_000),
    ('Libéria', 'Monróvia', 5_500_000),
    ('Líbia', 'Trípoli', 7_000_000),
    ('Liechtenstein', 'Vaduz', 40_000),
    ('Lituânia', 'Vilnius', 2_800_000),
    ('Luxemburgo', 'Luxemburgo', 700_000),
    ('Macedônia do Norte', 'Skopje', 2_000_000),
    ('Madagáscar', 'Antananarivo', 31_000_000),
    ('Malásia', 'Kuala Lumpur', 34_000_000),
    ('Malaui', 'Lilongwe', 21_000_000),
    ('Maldivas', 'Malé', 500_000),
    ('Mali', 'Bamako', 23_000_000),
    ('Malta', 'Valletta', 550_000),
    ('Marrocos', 'Rabat', 38_000_000),
    ('Marshall, Ilhas', 'Majuro', 60_000),
    ('Maurícia', 'Port Louis', 1_300_000),
    ('Mauritânia', 'Nouakchott', 5_000_000),
    ('México', 'Cidade do México', 130_000_000),
    ('Micronésia', 'Palikir', 100_000),
    ('Moçambique', 'Maputo', 34_000_000),
    ('Moldávia', 'Chisinau', 2_600_000),
    ('Mônaco', 'Mônaco', 40_000),
    ('Mongólia', 'Ulan Bator', 3_400_000),
    ('Montenegro', 'Podgorica', 600_000),
    ('Myanmar', 'Naypyidaw', 55_000_000),
    ('Namíbia', 'Windhoek', 2_600_000),
    ('Nauru', 'Yaren', 10_000),
    ('Nepal', 'Katmandu', 31_000_000),
    ('Nicarágua', 'Manágua', 7_000_000),
    ('Níger', 'Niamey', 27_000_000),
    ('Nigéria', 'Abuja', 230_000_000),
    ('Noruega', 'Oslo', 5_500_000),
    ('Nova Zelândia', 'Wellington', 5_300_000),
    ('Omã', 'Mascate', 5_500_000),
    ('Países Baixos', 'Amsterdã', 18_000_000),  # Nota: Haia é a sede do governo
    ('Palau', 'Ngerulmud', 20_000),
    ('Palestina', 'Ramallah', 5_500_000),  # Nota: Jerusalém Oriental é reivindicada; Ramallah é administrativa
    ('Panamá', 'Cidade do Panamá', 4_500_000),
    ('Papua Nova Guiné', 'Port Moresby', 10_000_000),
    ('Paquistão', 'Islamabad', 240_000_000),
    ('Paraguai', 'Assunção', 7_500_000),
    ('Peru', 'Lima', 34_000_000),
    ('Polônia', 'Varsóvia', 38_000_000),
    ('Portugal', 'Lisboa', 10_000_000),
    ('Quênia', 'Nairóbi', 56_000_000),
    ('Quirguistão', 'Bishkek', 7_000_000),
    ('Reino Unido', 'Londres', 68_000_000),
    ('República Centro-Africana', 'Bangui', 5_700_000),
    ('República Checa', 'Praga', 10_000_000),
    ('República Democrática do Congo', 'Kinshasa', 100_000_000),
    ('República Dominicana', 'Santo Domingo', 11_000_000),
    ('Romênia', 'Bucareste', 19_000_000),
    ('Ruanda', 'Kigali', 14_000_000),
    ('Rússia', 'Moscou', 146_000_000),
    ('Salomão, Ilhas', 'Honiara', 700_000),
    ('Samoa', 'Apia', 200_000),
    ('San Marino', 'San Marino', 30_000),
    ('Santa Lúcia', 'Castries', 200_000),
    ('São Cristóvão e Nevis', 'Basseterre', 50_000),
    ('São Tomé e Príncipe', 'São Tomé', 200_000),
    ('São Vicente e Granadinas', 'Kingstown', 100_000),
    ('Senegal', 'Dakar', 18_000_000),
    ('Serra Leoa', 'Freetown', 8_700_000),
    ('Sérvia', 'Belgrado', 6_600_000),
    ('Seychelles', 'Victoria', 100_000),
    ('Singapura', 'Singapura', 6_000_000),
    ('Síria', 'Damasco', 24_000_000),
    ('Somália', 'Mogadíscio', 18_000_000),
    ('Sri Lanka', 'Sri Jayawardenepura Kotte', 22_000_000),
    ('Sudão', 'Cartum', 48_000_000),
    ('Sudão do Sul', 'Juba', 11_000_000),
    ('Suécia', 'Estocolmo', 10_000_000),
    ('Suíça', 'Berna', 9_000_000),
    ('Suriname', 'Paramaribo', 600_000),
    ('Tadjiquistão', 'Dushanbe', 10_000_000),
    ('Tailândia', 'Bangcoc', 71_000_000),
    ('Taiwan', 'Taipei', 24_000_000),  # Incluído, mas com status disputado
    ('Tanzânia', 'Dodoma', 67_000_000),
    ('Timor-Leste', 'Díli', 1_400_000),
    ('Togo', 'Lomé', 9_000_000),
    ('Tonga', 'Nuku’alofa', 100_000),
    ('Trinidad e Tobago', 'Porto de Espanha', 1_500_000),
    ('Tunísia', 'Túnis', 12_000_000),
    ('Turcomenistão', 'Asgabate', 6_500_000),
    ('Turquia', 'Ancara', 85_000_000),
    ('Tuvalu', 'Funafuti', 10_000),
    ('Ucrânia', 'Kiev', 44_000_000),
    ('Uganda', 'Kampala', 50_000_000),
    ('Uruguai', 'Montevidéu', 3_500_000),
    ('Uzbequistão', 'Tashkent', 36_000_000),
    ('Vanuatu', 'Port Vila', 300_000),
    ('Vaticano', 'Cidade do Vaticano', 800),
    ('Venezuela', 'Caracas', 28_000_000),
    ('Vietnã', 'Hanói', 100_000_000),
    ('Zâmbia', 'Lusaka', 21_000_000),
    ('Zimbábue', 'Harare', 16_000_000),
]

# Definir as cores de preenchimento
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Vermelho para países
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarelo para capitais
fill_blue = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")    # Azul para população

# Inserir os países, capitais e população nas células com cores
for i, (pais, capital, populacao) in enumerate(paises_capitais_populacao, start=2):
    objWorksheet.cell(i, 1, pais).fill = fill_red        # Coluna 1: País (vermelho)
    objWorksheet.cell(i, 2, capital).fill = fill_yellow  # Coluna 2: Capital (amarelo)
    objWorksheet.cell(i, 3, populacao).fill = fill_blue  # Coluna 3: População (azul)

# Salvar a planilha
objWorkbook.save("World_Countries_Capitals_Population.xlsx")